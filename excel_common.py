"""Excel 文件读取和写入的通用工具。

这里放跨功能复用的逻辑，例如判断 Excel 文件类型、递归枚举 Excel 文件、
读取工作表名称、预览数据和导出结果。
"""

import os
from collections.abc import Iterator
from pathlib import Path
from typing import Iterable, Sequence
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import load_workbook
from xlsxwriter import Workbook


EXCEL_EXTENSIONS = {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".xlsb"}
EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384


class ExcelReadError(RuntimeError):
    """统一包装 Excel 读取、写入和路径校验时出现的错误。"""

    pass


def get_sheet_names(file_path: str | Path) -> list[str]:
    """根据文件扩展名选择最快的方式读取工作表名称。"""

    path = Path(file_path)
    if not path.exists():
        raise ExcelReadError(f"File does not exist: {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _get_xlsx_xml_sheet_names(path)
    if suffix == ".xls":
        return _get_xlrd_sheet_names(path)
    if suffix == ".xlsb":
        return _get_xlsb_sheet_names(path)

    raise ExcelReadError(f"Unsupported Excel file type: {path.suffix}")


def iter_excel_files(folder_path: str | Path, cancel_event: object | None = None) -> Iterator[Path]:
    """递归枚举文件夹内所有支持的 Excel 文件，并支持外部取消。"""

    folder = validate_excel_folder(folder_path)
    stack = [folder]
    while stack:
        if _is_cancelled(cancel_event):
            break

        current_folder = stack.pop()
        try:
            with os.scandir(current_folder) as scanner:
                entries = sorted(scanner, key=lambda entry: entry.path.lower())
        except OSError:
            continue

        child_folders: list[Path] = []
        for entry in entries:
            if _is_cancelled(cancel_event):
                break
            try:
                if entry.is_dir(follow_symlinks=False):
                    child_folders.append(Path(entry.path))
                    continue
                if not entry.is_file(follow_symlinks=False):
                    continue
            except OSError:
                continue

            path = Path(entry.path)
            if is_excel_file(path):
                yield path

        stack.extend(reversed(child_folders))


def validate_excel_folder(folder_path: str | Path) -> Path:
    """确认传入路径存在且是文件夹。"""

    folder = Path(folder_path)
    if not folder.exists():
        raise ExcelReadError(f"Folder does not exist: {folder}")
    if not folder.is_dir():
        raise ExcelReadError(f"Not a folder: {folder}")
    return folder


def is_excel_file(path: Path) -> bool:
    """判断路径是否是可处理的 Excel 文件，跳过 Excel 临时锁文件。"""

    if not path.is_file():
        return False
    if path.name.startswith("~$"):
        return False
    return path.suffix.lower() in EXCEL_EXTENSIONS


def _is_cancelled(cancel_event: object | None) -> bool:
    """兼容任意带 is_set 方法的取消事件对象。"""

    is_set = getattr(cancel_event, "is_set", None)
    return bool(callable(is_set) and is_set())


def clean_cell_text(value: object) -> str:
    """把单元格值转成去掉首尾空白的字符串。"""

    if value is None:
        return ""
    return str(value).strip()


def read_excel_preview(
    file_path: str | Path,
    sheet_name: str | None = None,
    max_rows: int = 20,
) -> list[list[object]]:
    """读取指定工作表前几行数据，用于需要快速预览的场景。"""

    path = Path(file_path)
    if not path.exists():
        raise ExcelReadError(f"File does not exist: {path}")

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc

    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows: list[list[object]] = []
        for row in worksheet.iter_rows(max_row=max_rows, values_only=True):
            rows.append(list(row))
        return rows
    finally:
        workbook.close()


def write_rows_to_excel(
    file_path: str | Path,
    rows: Iterable[Sequence[object]],
    sheet_name: str = "Sheet1",
) -> None:
    """把二维行数据写入新的 Excel 文件。"""

    path = Path(file_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook(str(path))
    worksheet = workbook.add_worksheet(sheet_name[:31] or "Sheet1")

    for row_index, row in enumerate(rows):
        for column_index, value in enumerate(row):
            worksheet.write(row_index, column_index, value)

    workbook.close()


def _get_xlsx_xml_sheet_names(path: Path) -> list[str]:
    """直接读取 xlsx 压缩包中的 workbook.xml，避免完整加载工作簿。"""

    try:
        with zipfile.ZipFile(path) as archive:
            with archive.open("xl/workbook.xml") as workbook_xml:
                root = ET.parse(workbook_xml).getroot()
    except KeyError as exc:
        raise ExcelReadError(f"Could not read workbook metadata: {path}") from exc
    except (OSError, zipfile.BadZipFile, ET.ParseError) as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc

    sheet_names: list[str] = []
    for element in root.iter():
        if _xml_local_name(element.tag) != "sheet":
            continue
        sheet_name = element.attrib.get("name")
        if sheet_name is not None:
            sheet_names.append(sheet_name)

    return sheet_names


def _xml_local_name(tag: str) -> str:
    """去掉 XML 命名空间，只保留标签本名。"""

    if "}" not in tag:
        return tag
    return tag.rsplit("}", 1)[1]


def _get_openpyxl_sheet_names(path: Path) -> list[str]:
    """使用 openpyxl 读取工作表名称的备用实现。"""

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc

    try:
        return list(workbook.sheetnames)
    finally:
        workbook.close()


def _get_xlrd_sheet_names(path: Path) -> list[str]:
    """读取老格式 .xls 文件的工作表名称。"""

    try:
        import xlrd
    except ImportError as exc:
        raise ExcelReadError("xlrd is required to read old .xls files") from exc

    try:
        workbook = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc

    try:
        return list(workbook.sheet_names())
    finally:
        workbook.release_resources()


def _get_xlsb_sheet_names(path: Path) -> list[str]:
    """读取二进制 .xlsb 文件的工作表名称。"""

    try:
        from pyxlsb import open_workbook
    except ImportError as exc:
        raise ExcelReadError("pyxlsb is required to read .xlsb files") from exc

    try:
        with open_workbook(str(path)) as workbook:
            return list(workbook.sheets)
    except Exception as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc
