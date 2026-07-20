"""工作簿、工作表和引用关系检索逻辑。

界面里的“工作表检索”和“查找引用”按钮主要调用这里的函数。
工作表名称会按文件大小和修改时间缓存，减少重复打开 Excel 的成本。
"""

import re
from dataclasses import dataclass
from pathlib import Path
from threading import Event, Lock

from openpyxl import load_workbook

from cache_store import get_app_directory, read_cache_data, update_cache_data
from excel_common import (
    EXCEL_MAX_ROWS,
    ExcelReadError,
    clean_cell_text,
    get_sheet_names,
    iter_excel_files,
)


@dataclass(frozen=True)
class ExcelWorkbookInfo:
    """一个工作簿的扫描结果。"""

    path: Path
    workbook_name: str
    sheet_names: list[str]
    error: str | None = None


@dataclass(frozen=True)
class ReferenceLookupConfig:
    """引用检索规则配置。

    sample_text 用来描述引用结构，table_name 和 field_name 标记其中哪一段
    是目标表名、哪一段是字段名。
    """

    sample_text: str = "FOREIGN:BBB.Id"
    reference_row_index: int = 6
    table_name: str = "BBB"
    field_name: str = "Id"
    field_row_index: int = 1


@dataclass(frozen=True)
class ExcelSheetReference:
    """从某个工作表引用行中解析出的单条引用关系。"""

    sheet_name: str
    reference_field_name: str
    source_field_name: str = ""


DEFAULT_REFERENCE_LOOKUP_CONFIG = ReferenceLookupConfig()
LEGACY_SHEET_NAME_CACHE_FILE = get_app_directory() / "worksheet_sheet_cache.json"
WORKSHEET_SHEET_CACHE_KEY = "worksheet_sheet_cache"
_sheet_name_cache: dict[str, dict[str, object]] | None = None
_sheet_name_cache_dirty = False
_sheet_name_cache_lock = Lock()


def scan_excel_workbooks(
    folder_path: str | Path,
    keyword: str = "",
    cancel_event: Event | None = None,
) -> list[ExcelWorkbookInfo]:
    """扫描目录内 Excel 工作簿，并按关键词过滤工作簿名或工作表名。"""

    normalized_keyword = keyword.strip().lower()
    results: list[ExcelWorkbookInfo] = []

    try:
        for path in iter_excel_files(folder_path, cancel_event=cancel_event):
            if _is_cancelled(cancel_event):
                break

            path_matches = _path_matches_keyword(path, normalized_keyword)
            try:
                sheet_names = _get_cached_sheet_names(path)
                error = None
            except ExcelReadError as exc:
                sheet_names = []
                error = str(exc)

            if _is_cancelled(cancel_event):
                break

            if error is not None:
                if not normalized_keyword or path_matches:
                    results.append(
                        ExcelWorkbookInfo(
                            path=path,
                            workbook_name=path.name,
                            sheet_names=[],
                            error=error,
                        )
                    )
                continue

            if not normalized_keyword or path_matches:
                matched_sheet_names = sheet_names
            else:
                matched_sheet_names = _filter_sheet_names(sheet_names, normalized_keyword)
                if not matched_sheet_names:
                    continue

            results.append(
                ExcelWorkbookInfo(
                    path=path,
                    workbook_name=path.name,
                    sheet_names=matched_sheet_names,
                    error=None,
                )
            )
    finally:
        _flush_sheet_name_cache()

    return results


def _is_cancelled(cancel_event: Event | None) -> bool:
    """判断后台扫描是否已被取消。"""

    return cancel_event is not None and cancel_event.is_set()


def _path_matches_keyword(path: Path, keyword: str) -> bool:
    """判断关键词是否命中文件名、文件主名或完整路径。"""

    if not keyword:
        return True

    return any(
        keyword in part.lower()
        for part in (
            path.name,
            path.stem,
            str(path),
        )
    )


def _filter_sheet_names(sheet_names: list[str], keyword: str) -> list[str]:
    """只保留名称中包含关键词的工作表。"""

    if not keyword:
        return sheet_names

    return [sheet_name for sheet_name in sheet_names if keyword in sheet_name.lower()]


def _get_cached_sheet_names(path: Path) -> list[str]:
    """读取工作表名称，并用文件大小和修改时间判断缓存是否有效。"""

    cache_key = str(path.resolve())
    size, mtime_ns = _get_file_cache_marker(path)

    with _sheet_name_cache_lock:
        cache = _load_sheet_name_cache_unlocked()
        cached_entry = cache.get(cache_key)
        cached_sheet_names = _read_cached_sheet_names(cached_entry, size, mtime_ns)
        if cached_sheet_names is not None:
            return cached_sheet_names

    sheet_names = get_sheet_names(path)

    with _sheet_name_cache_lock:
        cache = _load_sheet_name_cache_unlocked()
        cache[cache_key] = {
            "size": size,
            "mtime_ns": mtime_ns,
            "sheet_names": sheet_names,
        }
        global _sheet_name_cache_dirty
        _sheet_name_cache_dirty = True

    return sheet_names


def _get_file_cache_marker(path: Path) -> tuple[int, int]:
    """生成用于判断工作簿是否变化的缓存标记。"""

    try:
        file_stat = path.stat()
    except OSError as exc:
        raise ExcelReadError(f"Could not read file metadata: {path}") from exc
    return file_stat.st_size, file_stat.st_mtime_ns


def _load_sheet_name_cache_unlocked() -> dict[str, dict[str, object]]:
    """加载工作表名称缓存；调用者必须已经持有缓存锁。"""

    global _sheet_name_cache, _sheet_name_cache_dirty

    if _sheet_name_cache is not None:
        return _sheet_name_cache

    raw_data = read_cache_data()

    if not isinstance(raw_data, dict):
        _sheet_name_cache = {}
        return _sheet_name_cache

    raw_cache = raw_data.get(WORKSHEET_SHEET_CACHE_KEY)
    if not isinstance(raw_cache, dict):
        legacy_cache = _load_legacy_sheet_name_cache()
        if legacy_cache:
            _sheet_name_cache_dirty = True
        _sheet_name_cache = legacy_cache
        return _sheet_name_cache

    _sheet_name_cache = {
        str(cache_key): cache_entry
        for cache_key, cache_entry in raw_cache.items()
        if isinstance(cache_entry, dict)
    }
    return _sheet_name_cache


def _load_legacy_sheet_name_cache() -> dict[str, dict[str, object]]:
    """兼容旧版本单独的 worksheet_sheet_cache.json 缓存文件。"""

    try:
        import json

        raw_cache = json.loads(LEGACY_SHEET_NAME_CACHE_FILE.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}

    if not isinstance(raw_cache, dict):
        return {}

    return {
        str(cache_key): cache_entry
        for cache_key, cache_entry in raw_cache.items()
        if isinstance(cache_entry, dict)
    }


def _read_cached_sheet_names(
    cached_entry: dict[str, object] | None,
    size: int,
    mtime_ns: int,
) -> list[str] | None:
    """校验缓存条目，只有文件大小和修改时间完全一致才复用。"""

    if not isinstance(cached_entry, dict):
        return None
    if cached_entry.get("size") != size or cached_entry.get("mtime_ns") != mtime_ns:
        return None

    sheet_names = cached_entry.get("sheet_names")
    if not isinstance(sheet_names, list) or not all(isinstance(item, str) for item in sheet_names):
        return None

    return list(sheet_names)


def _flush_sheet_name_cache() -> None:
    """把内存中的工作表名称缓存写回统一缓存文件。"""

    global _sheet_name_cache_dirty

    with _sheet_name_cache_lock:
        if _sheet_name_cache is None or not _sheet_name_cache_dirty:
            return
        cache_payload = dict(_sheet_name_cache)
        _sheet_name_cache_dirty = False

    if not update_cache_data({WORKSHEET_SHEET_CACHE_KEY: cache_payload}):
        with _sheet_name_cache_lock:
            _sheet_name_cache_dirty = True


def find_sheet_references(
    file_path: str | Path,
    sheet_name: str,
    row_index: int = 6,
) -> list[str]:
    """查找指定行内引用到的工作表名称，保留给简化调用场景使用。"""

    config = ReferenceLookupConfig(reference_row_index=row_index)
    matches = find_sheet_reference_matches(file_path, sheet_name, config)
    references: list[str] = []
    for match in matches:
        if match.sheet_name and match.sheet_name not in references:
            references.append(match.sheet_name)
    return references


def find_sheet_reference_matches(
    file_path: str | Path,
    sheet_name: str,
    config: ReferenceLookupConfig | None = None,
    cancel_event: Event | None = None,
) -> list[ExcelSheetReference]:
    """按照配置解析指定工作表引用行里的所有引用关系。"""

    lookup_config = config or DEFAULT_REFERENCE_LOOKUP_CONFIG
    validate_reference_lookup_config(lookup_config)
    path = Path(file_path)
    if not path.exists():
        raise ExcelReadError(f"File does not exist: {path}")

    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _find_openpyxl_sheet_reference_matches(path, sheet_name, lookup_config, cancel_event)
    if suffix == ".xls":
        return _find_xlrd_sheet_reference_matches(path, sheet_name, lookup_config, cancel_event)
    if suffix == ".xlsb":
        return _find_xlsb_sheet_reference_matches(path, sheet_name, lookup_config, cancel_event)

    raise ExcelReadError(f"Unsupported Excel file type: {path.suffix}")


def validate_reference_lookup_config(config: ReferenceLookupConfig) -> None:
    """通过构建正则表达式来验证引用配置是否完整、可解析。"""

    if config.reference_row_index > EXCEL_MAX_ROWS:
        raise ExcelReadError(f"引用行不能大于 {EXCEL_MAX_ROWS}")
    if config.field_row_index > EXCEL_MAX_ROWS:
        raise ExcelReadError(f"引用字段行不能大于 {EXCEL_MAX_ROWS}")
    _build_reference_pattern(config)


def _find_openpyxl_sheet_reference_matches(
    path: Path,
    sheet_name: str,
    config: ReferenceLookupConfig,
    cancel_event: Event | None,
) -> list[ExcelSheetReference]:
    """读取 xlsx/xlsm 等 openpyxl 支持格式的引用关系。"""

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc

    try:
        if _is_cancelled(cancel_event):
            return []
        if sheet_name not in workbook.sheetnames:
            raise ExcelReadError(f"Sheet does not exist: {sheet_name}")

        worksheet = workbook[sheet_name]
        pattern = _build_reference_pattern(config)
        source_fields = _get_openpyxl_row_values(worksheet, config.field_row_index)
        references: list[ExcelSheetReference] = []
        for row in worksheet.iter_rows(
            min_row=config.reference_row_index,
            max_row=config.reference_row_index,
        ):
            for column_index, cell in enumerate(row, start=1):
                if _is_cancelled(cancel_event):
                    return references
                _add_unique_reference_matches(
                    references,
                    cell.value,
                    source_fields.get(column_index, ""),
                    pattern,
                )
        return references
    finally:
        workbook.close()


def _find_xlrd_sheet_reference_matches(
    path: Path,
    sheet_name: str,
    config: ReferenceLookupConfig,
    cancel_event: Event | None,
) -> list[ExcelSheetReference]:
    """读取老格式 .xls 文件的引用关系。"""

    try:
        import xlrd
    except ImportError as exc:
        raise ExcelReadError("xlrd is required to read old .xls files") from exc

    try:
        workbook = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc

    try:
        try:
            worksheet = workbook.sheet_by_name(sheet_name)
        except xlrd.XLRDError as exc:
            raise ExcelReadError(f"Sheet does not exist: {sheet_name}") from exc

        row_offset = config.reference_row_index - 1
        references: list[ExcelSheetReference] = []
        if worksheet.nrows <= row_offset:
            return references

        pattern = _build_reference_pattern(config)
        source_fields = _get_xlrd_row_values(worksheet, config.field_row_index)
        for column_offset in range(worksheet.ncols):
            if _is_cancelled(cancel_event):
                return references
            _add_unique_reference_matches(
                references,
                worksheet.cell_value(row_offset, column_offset),
                source_fields.get(column_offset + 1, ""),
                pattern,
            )
        return references
    finally:
        workbook.release_resources()


def _find_xlsb_sheet_reference_matches(
    path: Path,
    sheet_name: str,
    config: ReferenceLookupConfig,
    cancel_event: Event | None,
) -> list[ExcelSheetReference]:
    """读取 .xlsb 文件的引用关系，只遍历配置里需要的行。"""

    try:
        from pyxlsb import open_workbook
    except ImportError as exc:
        raise ExcelReadError("pyxlsb is required to read .xlsb files") from exc

    try:
        with open_workbook(str(path)) as workbook:
            if sheet_name not in workbook.sheets:
                raise ExcelReadError(f"Sheet does not exist: {sheet_name}")

            pattern = _build_reference_pattern(config)
            references: list[ExcelSheetReference] = []
            source_fields: dict[int, str] = {}
            reference_values: list[tuple[int, object]] = []
            last_needed_row = max(config.reference_row_index, config.field_row_index)
            with workbook.get_sheet(sheet_name) as worksheet:
                for current_row_index, row in enumerate(worksheet.rows(), start=1):
                    if _is_cancelled(cancel_event):
                        return references
                    if current_row_index == config.field_row_index:
                        for column_index, cell in enumerate(row, start=1):
                            source_fields[column_index] = clean_cell_text(cell.v)
                    if current_row_index == config.reference_row_index:
                        reference_values = [
                            (column_index, cell.v)
                            for column_index, cell in enumerate(row, start=1)
                        ]
                    if current_row_index >= last_needed_row:
                        break

            for column_index, value in reference_values:
                _add_unique_reference_matches(
                    references,
                    value,
                    source_fields.get(column_index, ""),
                    pattern,
                )
            return references
    except ExcelReadError:
        raise
    except Exception as exc:
        raise ExcelReadError(f"Could not open Excel file: {path}") from exc


def _get_openpyxl_row_values(worksheet: object, row_index: int) -> dict[int, str]:
    """读取 openpyxl 工作表指定行的值，并按 1 起始列号保存。"""

    values: dict[int, str] = {}
    for row in worksheet.iter_rows(min_row=row_index, max_row=row_index):
        for column_index, cell in enumerate(row, start=1):
            values[column_index] = clean_cell_text(cell.value)
    return values


def _get_xlrd_row_values(worksheet: object, row_index: int) -> dict[int, str]:
    """读取 xlrd 工作表指定行的值，并按 1 起始列号保存。"""

    row_offset = row_index - 1
    if worksheet.nrows <= row_offset:
        return {}

    return {
        column_offset + 1: clean_cell_text(worksheet.cell_value(row_offset, column_offset))
        for column_offset in range(worksheet.ncols)
    }


def _add_unique_reference_matches(
    references: list[ExcelSheetReference],
    value: object,
    source_field_name: str,
    pattern: re.Pattern[str],
) -> None:
    """从单元格文本中提取引用关系，并避免重复加入结果。"""

    if value is None:
        return

    for match in pattern.finditer(str(value)):
        sheet_name = match.group("table").strip()
        reference_field_name = match.group("field").strip()
        if not sheet_name:
            continue

        reference = ExcelSheetReference(
            sheet_name=sheet_name,
            reference_field_name=reference_field_name,
            source_field_name=source_field_name.strip(),
        )
        if reference not in references:
            references.append(reference)


def _build_reference_pattern(config: ReferenceLookupConfig) -> re.Pattern[str]:
    """把用户配置的引用样例转换成可提取表名和字段名的正则。"""

    sample_text = config.sample_text.strip()
    table_name = config.table_name.strip()
    field_name = config.field_name.strip()

    if not sample_text:
        raise ExcelReadError("引用样例不能为空")
    if not table_name:
        raise ExcelReadError("引用表格名称不能为空")
    if not field_name:
        raise ExcelReadError("引用字段不能为空")
    if config.reference_row_index < 1:
        raise ExcelReadError("引用行必须大于等于 1")
    if config.field_row_index < 1:
        raise ExcelReadError("引用字段行必须大于等于 1")

    table_start = sample_text.find(table_name)
    field_start = sample_text.find(field_name)
    if table_start < 0:
        raise ExcelReadError("引用样例中找不到填写的引用表格名称")
    if field_start < 0:
        raise ExcelReadError("引用样例中找不到填写的引用字段")

    spans = [
        ("table", table_start, len(table_name)),
        ("field", field_start, len(field_name)),
    ]
    spans.sort(key=lambda item: item[1])
    first_end = spans[0][1] + spans[0][2]
    if first_end > spans[1][1]:
        raise ExcelReadError("引用表格名称和引用字段在样例中不能重叠")

    pattern_parts: list[str] = []
    cursor = 0
    for index, (group_name, start, length) in enumerate(spans):
        pattern_parts.append(re.escape(sample_text[cursor:start]))
        next_start = spans[index + 1][1] if index + 1 < len(spans) else len(sample_text)
        next_literal = sample_text[start + length:next_start]
        # 如果后面还有固定文本，用非贪婪匹配；末尾字段则遇到空白或常见分隔符停止。
        pattern_parts.append(f"(?P<{group_name}>.+?)" if next_literal else f"(?P<{group_name}>[^\\s,;，；]+)")
        cursor = start + length
    pattern_parts.append(re.escape(sample_text[cursor:]))

    return re.compile("".join(pattern_parts), re.IGNORECASE)
